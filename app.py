from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import mm
import io
import zipfile
import re
import os
import traceback
from urllib.parse import quote

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# 注册中文字体（使用 reportlab 内置的中文字体）
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    FONT_NAME = 'STSong-Light'
    print("使用中文字体: STSong-Light")
except:
    FONT_NAME = 'Helvetica'
    print("警告: 使用默认字体，中文可能显示为方框")

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/process', methods=['POST'])
def process_file():
    try:
        print("=" * 50)
        print("收到处理请求")
        
        if 'file' not in request.files:
            return jsonify({'error': '未上传文件'}), 400
        
        file = request.files['file']
        project_id = request.form.get('project_id', '').strip()
        
        print(f"项目编号: {project_id}")
        print(f"文件名: {file.filename}")
        
        if not project_id:
            return jsonify({'error': '请输入项目编号'}), 400
        
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        print(f"Excel 读取成功，共 {ws.max_row} 行")
        
        # 提取元数据
        drug_name = extract_drug_name(ws)
        date_range = extract_date_range(ws)
        
        if not drug_name:
            drug_name = "未知药品"
        if not date_range:
            date_range = "未知日期"
        
        print(f"提取信息 - 药品: {drug_name}, 日期: {date_range}")
        
        # 清理文件名（保留中文）
        drug_name_clean = re.sub(r'[\\/:*?"<>|]', '_', drug_name)
        date_range_clean = re.sub(r'[\\/:*?"<>|]', '_', date_range)
        
        # 查找项目编号列
        project_col, data_start_row = find_project_column(ws)
        
        if project_col is None:
            return jsonify({'error': '未找到项目编号列'}), 400
        
        # 分类数据
        matching_rows = []
        non_matching_rows = []
        
        for row_idx in range(data_start_row + 1, ws.max_row + 1):
            cell_value = str(ws.cell(row_idx, project_col).value or '').strip()
            if cell_value == project_id:
                matching_rows.append(row_idx)
            elif cell_value:
                non_matching_rows.append(row_idx)
        
        print(f"本项目: {len(matching_rows)} 条, 非本项目: {len(non_matching_rows)} 条")
        
        # 生成 PDF 文件
        pdf_files = []
        
        if matching_rows:
            print("生成本项目 PDF...")
            pdf_buffer = create_pdf(ws, data_start_row, matching_rows)
            filename = f'本项目外院SUSAR_{drug_name_clean}_{date_range_clean}.pdf'
            pdf_files.append((filename, pdf_buffer))
        
        if non_matching_rows:
            print("生成非本项目 PDF...")
            pdf_buffer = create_pdf(ws, data_start_row, non_matching_rows)
            filename = f'非本项目外院SUSAR_{drug_name_clean}_{date_range_clean}.pdf'
            pdf_files.append((filename, pdf_buffer))
        
        if not pdf_files:
            return jsonify({'error': '没有找到任何数据'}), 400
        
        # 如果只有一个文件，直接返回
        if len(pdf_files) == 1:
            filename, buffer = pdf_files[0]
            print(f"返回文件: {filename}")
            
            # 使用 RFC 5987 编码文件名
            encoded_filename = quote(filename)
            
            response = send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
            response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
            return response
        
        # 多个文件打包成 ZIP
        print("打包为 ZIP...")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, buffer in pdf_files:
                zip_file.writestr(filename, buffer.getvalue())
        
        zip_buffer.seek(0)
        zip_filename = f'SUSAR_{drug_name_clean}_{date_range_clean}.zip'
        
        encoded_filename = quote(zip_filename)
        
        response = send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response
        
    except Exception as e:
        print("错误:", str(e))
        print(traceback.format_exc())
        return jsonify({'error': f'处理失败: {str(e)}'}), 500

def extract_drug_name(ws):
    for row in range(1, min(11, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell_value = str(ws.cell(row, col).value or '')
            if 'Investigational Drug' in cell_value or '试验药物' in cell_value:
                match = re.search(r'(?:Investigational Drug|试验药物)[:\s：]+(.+)', cell_value, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
                if col + 1 <= ws.max_column:
                    next_val = str(ws.cell(row, col + 1).value or '').strip()
                    if next_val and len(next_val) > 0:
                        return next_val
                if row + 1 <= ws.max_row:
                    below_val = str(ws.cell(row + 1, col).value or '').strip()
                    if below_val and len(below_val) > 0:
                        return below_val
    return None

def extract_date_range(ws):
    for row in range(1, min(11, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell_value = str(ws.cell(row, col).value or '')
            if any(keyword in cell_value for keyword in ['传输数据区间', 'Data Transfer Period', '数据区间', '日期区间']):
                match = re.search(r'[:\s：]+(.+)', cell_value)
                if match:
                    extracted = match.group(1).strip()
                    if len(extracted) > 3:
                        return extracted
                if col + 1 <= ws.max_column:
                    next_val = str(ws.cell(row, col + 1).value or '').strip()
                    if next_val and len(next_val) > 0:
                        return next_val
                if row + 1 <= ws.max_row:
                    below_val = str(ws.cell(row + 1, col).value or '').strip()
                    if below_val and len(below_val) > 0:
                        return below_val
    return None

def find_project_column(ws):
    for row in range(1, min(11, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 30)):
            cell_value = str(ws.cell(row, col).value or '').lower()
            if any(keyword in cell_value for keyword in ['study', '项目', 'study id', 'protocol', '编号', '方案']):
                return col, row
    return None, None

def create_pdf(ws, data_start_row, keep_rows):
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=5*mm,
        rightMargin=5*mm,
        topMargin=5*mm,
        bottomMargin=5*mm
    )
    
    table_data = []
    
    # 添加头部行
    for row in range(1, data_start_row + 1):
        row_data = []
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row, col)
            value = str(cell.value or '').strip()
            # 限制长度避免表格过宽
            if len(value) > 40:
                value = value[:37] + '...'
            row_data.append(value)
        table_data.append(row_data)
    
    # 添加数据行
    for row_idx in keep_rows:
        row_data = []
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row_idx, col)
            value = str(cell.value or '').strip()
            if len(value) > 40:
                value = value[:37] + '...'
            row_data.append(value)
        table_data.append(row_data)
    
    # 创建表格
    table = Table(table_data)
    
    # 设置表格样式
    style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 5.5),
        ('FONTSIZE', (0, 0), (-1, data_start_row - 1), 7),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, data_start_row - 1), colors.HexColor('#E8E8E8')),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])
    
    table.setStyle(style)
    
    doc.build([table])
    buffer.seek(0)
    return buffer

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
